import csv
import os
from datetime import datetime
from app.models import Student
import openpyxl
from openpyxl.styles import Font, Alignment

class ExportService:

    def _get_students(self, filters):
        """Récupère les élèves selon les filtres"""
        query = Student.query

        if filters.get('ville'):
            query = query.filter_by(ville=filters['ville'])
        if filters.get('ecole'):
            query = query.filter_by(ecole=filters['ecole'])
        if filters.get('status'):
            query = query.filter_by(status=filters['status'])

        return query.order_by(Student.ville, Student.ecole, Student.nom).all()

    def export_csv(self, filters=None):
        """Exporte en CSV"""
        filters = filters or {}
        students = self._get_students(filters)

        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        export_dir = os.path.join('data', 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            fieldnames = ['ville', 'ecole', 'classe', 'nom', 'prenom', 'age',
                         'acuite_og', 'acuite_od', 'sph_og', 'cyl_og', 'axe_og',
                         'sph_od', 'cyl_od', 'axe_od', 'ecart_pupillaire',
                         'prise_en_charge', 'observations', 'status']

            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for student in students:
                writer.writerow({
                    'ville': student.ville,
                    'ecole': student.ecole,
                    'classe': student.classe,
                    'nom': student.nom,
                    'prenom': student.prenom,
                    'age': student.age,
                    'acuite_og': student.acuite_og,
                    'acuite_od': student.acuite_od,
                    'sph_og': student.sph_og,
                    'cyl_og': student.cyl_og,
                    'axe_og': student.axe_og,
                    'sph_od': student.sph_od,
                    'cyl_od': student.cyl_od,
                    'axe_od': student.axe_od,
                    'ecart_pupillaire': student.ecart_pupillaire,
                    'prise_en_charge': student.prise_en_charge,
                    'observations': student.observations,
                    'status': student.status
                })

        return filepath

    def export_excel(self, filters=None):
        """Exporte en Excel"""
        filters = filters or {}
        students = self._get_students(filters)

        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_dir = os.path.join('data', 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Élèves"

        # En-têtes
        headers = ['Ville', 'École', 'Classe', 'Nom', 'Prénom', 'Âge',
                   'Acuité OG', 'Acuité OD', 'Sph OG', 'Cyl OG', 'Axe OG',
                   'Sph OD', 'Cyl OD', 'Axe OD', 'EP', 'Prise en charge',
                   'Observations', 'Statut']

        ws.append(headers)

        # Style des en-têtes
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Données
        for student in students:
            ws.append([
                student.ville, student.ecole, student.classe,
                student.nom, student.prenom, student.age,
                student.acuite_og, student.acuite_od,
                student.sph_og, student.cyl_og, student.axe_og,
                student.sph_od, student.cyl_od, student.axe_od,
                student.ecart_pupillaire, student.prise_en_charge,
                student.observations, student.status
            ])

        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        return filepath
